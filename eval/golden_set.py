"""
Golden intake set — ~150 stratified fixtures derived from the MASA use-case dataset.

Source: MASA_Use_Case_Coverage_Analysis.xlsx (2,848 rows, repo root)
Method: Stratified sample across Primary Need × PRD Fit; reproducible via random.seed(42).

Each fixture is a dict where:
  - Plain keys are valid IntakeSubmission fields
  - __ prefixed keys are ground-truth labels (stripped before constructing IntakeSubmission)

Population caveat (PRD §14): Use cases were mined from CFPB complaints and Reddit
(general population, collections-heavy). MASA members skew 55+, Medicare/MA,
ambulance-centric. This set demonstrates the engine on complaint-derived cases only.
"""

from __future__ import annotations

import logging
import random
from pathlib import Path

logger = logging.getLogger(__name__)

DATASET_PATH = Path(__file__).parent.parent / "MASA_Use_Case_Coverage_Analysis.xlsx"

# Primary Need → triage label mapping (PRD §4, §15)
# primary_workflow values reflect Axis-1 routing from triage/engine.py
_NEED_MAP: dict[str, dict] = {
    "itemized_bill_guidance":             {"problem_type": "billing_error",             "primary_workflow": "workflow_1"},
    "collection_abuse_only":              {"problem_type": "collections_credit_impact", "primary_workflow": "workflow_4", "in_collections": True},
    "cost_estimate_dispute":              {"problem_type": "billing_error",             "primary_workflow": "workflow_1"},
    "duplicate_billing":                  {"problem_type": "billing_error",             "primary_workflow": "workflow_1"},
    "plan_benefit_lookup":                {"problem_type": "billing_error",             "primary_workflow": "workflow_1"},
    "code_decode_eob":                    {"problem_type": "billing_error",             "primary_workflow": "workflow_1",
                                           "codes_present": [{"code_type": "ICD10CM", "code": "I10"}]},
    "code_decode":                        {"problem_type": "billing_error",             "primary_workflow": "workflow_1",
                                           "codes_present": [{"code_type": "HCPCS", "code": "A0425"}]},
    "transparency_unspecified":           {"problem_type": "billing_error",             "primary_workflow": "workflow_1"},
    "transparency_general":               {"problem_type": "billing_error",             "primary_workflow": "workflow_1"},
    "coordination_of_benefits":           {"problem_type": "billing_error",             "primary_workflow": "workflow_1"},
    "medical_necessity_denial":           {"problem_type": "clean_denial",              "primary_workflow": "workflow_1", "denial_present": True},
    "appeal_pathway":                     {"problem_type": "clean_denial",              "primary_workflow": "workflow_1", "denial_present": True},
    "nsa_surprise_billing":               {"problem_type": "surprise_out_of_network",   "primary_workflow": "workflow_2"},
    "nsa_general":                        {"problem_type": "surprise_out_of_network",   "primary_workflow": "workflow_2"},
    "network_verification":               {"problem_type": "surprise_out_of_network",   "primary_workflow": "workflow_2"},
    "ambulance_general":                  {"problem_type": "surprise_out_of_network",   "primary_workflow": "workflow_2",
                                           "ambulance_involved": True, "ambulance_type": "ground"},
    "ambulance_advocacy":                 {"problem_type": "surprise_out_of_network",   "primary_workflow": "workflow_2",
                                           "ambulance_involved": True, "ambulance_type": "ground"},
    "collection_with_underlying_billing": {"problem_type": "collections_credit_impact", "primary_workflow": "workflow_4", "in_collections": True},
    "plan_benefit_with_collection":       {"problem_type": "collections_credit_impact", "primary_workflow": "workflow_4"},
    "medicare_coverage":                  {"problem_type": "billing_error",             "primary_workflow": "workflow_1",
                                           "insurance_situation": "medicare_only"},
    "medicare_general":                   {"problem_type": "billing_error",             "primary_workflow": "workflow_1",
                                           "insurance_situation": "medicare_only"},
    # These map to workflow_5 (expected); triage engine routes billing_error → workflow_1,
    # so these will be counted as triage mismatches (contributing to accuracy <100%). Correct per PRD.
    "affordability_only":                 {"problem_type": "billing_error",             "primary_workflow": "workflow_5"},
    "other_unclassified":                 {"problem_type": "billing_error",             "primary_workflow": "workflow_5"},
    "identity_fraud_other":               {"problem_type": "billing_error",             "primary_workflow": "workflow_5"},
    "identity_fraud_provider":            {"problem_type": "billing_error",             "primary_workflow": "workflow_5"},
    "disability_cobra":                   {"problem_type": "billing_error",             "primary_workflow": "workflow_5"},
}

_DEFAULT_TRIAGE: dict = {"problem_type": "billing_error", "primary_workflow": "workflow_1"}

_ELIGIBLE_PRD_FIT = {"STRONG", "MODERATE", "PARTIAL"}


def _load_xlsx() -> list[dict]:
    """Load all rows from the dataset. Returns [] if file is missing."""
    if not DATASET_PATH.exists():
        logger.warning("Dataset not found at %s — GOLDEN_FIXTURES will be empty.", DATASET_PATH)
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(DATASET_PATH), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(next(rows_iter))]
        data = []
        for row in rows_iter:
            data.append(dict(zip(headers, row)))
        wb.close()
        return data
    except Exception as exc:
        logger.error("Failed to load xlsx: %s", exc)
        return []


def _row_to_fixture(row: dict, fixture_id: str) -> dict:
    """Convert one dataset row to a runnable intake fixture with ground-truth labels."""
    primary_need = (row.get("Primary Need") or "").strip()
    triage_map = _NEED_MAP.get(primary_need, _DEFAULT_TRIAGE)

    base: dict = {
        "insurance_situation": triage_map.get("insurance_situation", "commercial_employer"),
        "state": "FL",
        "problem_type": triage_map["problem_type"],
        "denial_present": triage_map.get("denial_present", False),
        "in_collections": triage_map.get("in_collections", False),
        "ambulance_involved": triage_map.get("ambulance_involved", False),
        "advocacy_capacity": "needs_hand_holding",
    }
    if triage_map.get("ambulance_type"):
        base["ambulance_type"] = triage_map["ambulance_type"]
    if triage_map.get("codes_present"):
        base["codes_present"] = triage_map["codes_present"]

    labels: dict = {
        "__fixture_id": fixture_id,
        "__primary_need": primary_need,
        "__prd_fit": (row.get("PRD Fit") or "").strip(),
        "__source": (row.get("Source") or "").strip(),
        "__case_title": str(row.get("Title") or "")[:200],
        "__pain_point_1": str(row.get("Pain Point 1") or "")[:500],
        "__expected_problem_type": triage_map["problem_type"],
        "__expected_insurance_situation": base["insurance_situation"],
        "__expected_primary_workflow": triage_map["primary_workflow"],
    }

    return {**base, **labels}


def _stratified_sample(rows: list[dict], n: int = 150, max_per_need: int = 10) -> list[dict]:
    """
    Stratified sample across Primary Need, filtered by PRD Fit.
    Reproducible: random.seed(42).
    """
    eligible = [
        r for r in rows
        if (r.get("PRD Fit") or "").strip() in _ELIGIBLE_PRD_FIT
    ]

    groups: dict[str, list[dict]] = {}
    for row in eligible:
        need = (row.get("Primary Need") or "unknown").strip()
        groups.setdefault(need, []).append(row)

    rng = random.Random(42)
    selected: list[dict] = []
    for need, group in sorted(groups.items()):
        k = min(max_per_need, len(group))
        selected.extend(rng.sample(group, k))

    if len(selected) > n:
        rng2 = random.Random(42)
        selected = rng2.sample(selected, n)

    return selected


def _build_golden_fixtures() -> list[dict]:
    rows = _load_xlsx()
    if not rows:
        return []
    sample = _stratified_sample(rows, n=150)
    fixtures = []
    for i, row in enumerate(sample):
        fixture_id = f"GS-{i + 1:04d}"
        fixtures.append(_row_to_fixture(row, fixture_id))
    logger.info("Golden set: %d fixtures from %d eligible rows.", len(fixtures), len(rows))
    return fixtures


# Module-level constant — built at import time
GOLDEN_FIXTURES: list[dict] = _build_golden_fixtures()
